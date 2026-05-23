"""
One-shot parser for "Энциклопедия алкоголя.xlsx".

  python3 scripts/parse-encyclopedia-xlsx.py \
      "Энциклопедия алкоголя.xlsx" \
      backend/data/encyclopedia_seed.json

Each sheet = one category. The "Выведенные" sheet is flagged
is_archived (still imported, but the public page can choose to hide it).

Columns are looked up by header name when a header row is present;
sheets without headers (e.g. "Выведенные") fall back to the canonical
column order: Название, %, $, Flavour, Бренд/Страна, Особенности,
Коктейль, Занимательный факт.
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

TRANSLIT = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
    'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
}

ARCHIVE_SHEETS = {'Выведенные'}

# Canonical header → field name. Match by normalised (lowercased, stripped) text.
HEADER_MAP = {
    'название':         'name',
    '%':                'abv',
    '$':                'price',
    'flavour':          'flavour',
    'бренд/страна':     'brand_country',
    'особенности':      'features',
    'коктейль':         'cocktail_pairings',
    'занимательный факт': 'fact',
}

FALLBACK_COLUMNS = [
    'name', 'abv', 'price', 'flavour', 'brand_country',
    'features', 'cocktail_pairings', 'fact',
]


def slugify(s: str, fallback: str = 'item') -> str:
    s = (s or '').strip().lower()
    if not s:
        return fallback
    out = []
    for ch in s:
        if ch in TRANSLIT:
            out.append(TRANSLIT[ch])
        elif ch.isalnum():
            out.append(ch)
        elif ch in (' ', '-', '_'):
            out.append('-')
    slug = ''.join(out)
    slug = re.sub(r'-+', '-', slug).strip('-')
    return slug or fallback


def stringify(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


URL_RE = re.compile(r'https?://\S+')


def extract_url_from_name(name: str | None) -> tuple[str | None, str | None]:
    """Returns (clean_name, url) — URL is moved out, name is whitespace-trimmed."""
    if not name:
        return name, None
    m = URL_RE.search(name)
    if not m:
        return name, None
    url = m.group(0).rstrip('.,;')
    clean = URL_RE.sub('', name)
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean or name, url


# Ordered most-specific → least-specific. Multi-word prefixes catch
# "Страна производителя" before single "Страна".
LINE_PREFIXES = [
    ('страна производителя', 'country'),
    ('производитель',        'brand'),
    ('бренд / серия',        'brand'),
    ('бренд/серия',          'brand'),
    ('бренд',                'brand'),
    ('страна',               'country'),
    ('регион',               'country'),
]

# Standalone country-only lines like "Мексика, Халиско" or "Шотландия".
RU_COUNTRY_HINT = re.compile(
    r'^(Россия|Мексика|США|Шотландия|Ирландия|Англия|Великобритания|Франция|Италия|'
    r'Испания|Германия|Куба|Перу|Чили|Бразилия|Япония|Корея|Грузия|Молдова|Армения|'
    r'Чехия|Польша|Венгрия|Финляндия|Швеция|Нидерланды|Австралия|Австрия)\b',
    re.IGNORECASE,
)

# A single line counts as structured only if it's short — otherwise it's
# prose ("Компания X была основана в...") and we leave it as notes.
MAX_STRUCTURED_LINE = 120


def split_brand_country(text: str | None) -> tuple[str | None, str | None, str | None]:
    """Conservative extraction: only short, clearly structured lines turn
    into brand / country. Long prose stays in `leftover_notes`."""
    if not text:
        return None, None, None
    brand_lines: list[str] = []
    country_lines: list[str] = []
    other_lines: list[str] = []

    for raw in text.split('\n'):
        line = raw.strip()
        if not line:
            continue
        low = line.lower().lstrip()

        handled = False
        if len(line) <= MAX_STRUCTURED_LINE:
            for prefix, field in LINE_PREFIXES:
                if low.startswith(prefix):
                    rest = line[len(prefix):].lstrip(' :\t-/').strip()
                    if rest:
                        (brand_lines if field == 'brand' else country_lines).append(rest)
                    handled = True
                    break
            if not handled and RU_COUNTRY_HINT.match(line) and len(line) < 80:
                country_lines.append(line)
                handled = True

        if not handled:
            other_lines.append(line)

    # De-duplicate and join
    def clean_join(parts, sep):
        seen = []
        for p in parts:
            p = p.strip(' ,.;')
            if p and p not in seen:
                seen.append(p)
        return sep.join(seen) or None

    brand = clean_join(brand_lines, '; ')
    country = clean_join(country_lines, ', ')
    leftover = '\n'.join(other_lines) or None
    return brand, country, leftover


def detect_header_row(ws) -> tuple[int, dict[int, str]] | tuple[None, None]:
    """Find the row index (1-based) where col A contains 'Название'.
    Returns (row_index, {col_index: field_name})."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        a = stringify(row[0])
        if a and 'название' in a.lower():
            mapping = {}
            for i, cell in enumerate(row):
                key = stringify(cell)
                if not key:
                    continue
                k_norm = key.lower().strip()
                if k_norm in HEADER_MAP:
                    mapping[i] = HEADER_MAP[k_norm]
            return row_idx, mapping
    return None, None


def parse(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    categories = []
    spirits = []
    seen_spirit_slugs: dict[str, int] = {}

    for cat_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        cat_slug = slugify(sheet_name)
        is_archived = sheet_name in ARCHIVE_SHEETS
        categories.append({
            'slug': cat_slug,
            'label': sheet_name.strip(),
            'sort_order': cat_idx,
            'is_archived': is_archived,
        })

        header_row, mapping = detect_header_row(ws)
        start_row = (header_row + 1) if header_row else 2  # for headerless sheets, skip the blank row 1
        if not mapping:
            # Use canonical positional mapping
            mapping = {i: f for i, f in enumerate(FALLBACK_COLUMNS)}

        sort_in_cat = 0
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            name = stringify(row[0])
            if not name:
                continue
            data = {'name': name}
            for col_idx, field in mapping.items():
                if col_idx == 0 or col_idx >= len(row):
                    continue
                v = stringify(row[col_idx])
                if v is not None:
                    data[field] = v

            clean_name, url_in_name = extract_url_from_name(data.get('name'))
            brand, country, leftover = split_brand_country(data.get('brand_country'))

            slug = slugify(clean_name or '', fallback=f'spirit-{len(spirits)+1}')[:70]
            n = seen_spirit_slugs.get(slug, 0)
            if n:
                slug = f'{slug}-{n+1}'
            seen_spirit_slugs[slug] = n + 1

            spirits.append({
                'slug': slug[:80],
                'category_slug': cat_slug,
                'name': clean_name,
                'source_url': url_in_name,
                'abv': data.get('abv'),
                'price': data.get('price'),
                'flavour': data.get('flavour'),
                'brand': brand,
                'country': country,
                'brand_country': leftover,        # only the lines we couldn't classify
                'features': data.get('features'),
                'cocktail_pairings': data.get('cocktail_pairings'),
                'fact': data.get('fact'),
                'sort_order': sort_in_cat,
            })
            sort_in_cat += 1

    return {'categories': categories, 'spirits': spirits}


def main():
    if len(sys.argv) < 3:
        print("usage: parse-encyclopedia-xlsx.py <input.xlsx> <output.json>", file=sys.stderr)
        sys.exit(2)
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2])
    data = parse(src)
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"  categories: {len(data['categories'])} ({sum(1 for c in data['categories'] if c['is_archived'])} archived)")
    print(f"  spirits:    {len(data['spirits'])}")
    by_cat = {}
    for s in data['spirits']:
        by_cat[s['category_slug']] = by_cat.get(s['category_slug'], 0) + 1
    for c in data['categories']:
        flag = ' (archived)' if c['is_archived'] else ''
        print(f"    {c['slug']:18} {c['label']:18} = {by_cat.get(c['slug'], 0)} items{flag}")
    print(f"  wrote: {dst}")


if __name__ == '__main__':
    main()
