"""
One-shot parser for "Меню кухня КОЛЛЕКТИВ.xlsx".

Reads the xlsx (single sheet, free-form layout) and emits a JSON
suitable for backend/seed.py:

  python3 scripts/parse-kitchen-xlsx.py \
      "Меню кухня КОЛЛЕКТИВ.xlsx" \
      backend/data/kitchen_seed.json

Layout assumption (from inspecting the workbook):
  Row 1: title  ("Меню кухня КОЛЛЕКТИВ")
  Row 2: header (Наименование | description | тайминг(мин) | выход(гр) | пищевая ценность | Сервировка)
  Rows 3+: alternating category-header rows (only col A populated) and dish rows
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

TRANSLIT = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
    'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
}


def slugify(s: str, fallback: str = 'item') -> str:
    s = (s or '').strip().lower()
    if not s:
        return fallback
    out = []
    for ch in s:
        if ch in TRANSLIT:
            out.append(TRANSLIT[ch])
        elif ch.isalnum():
            out.append(ch)
        elif ch in (' ', '-', '_'):
            out.append('-')
        # else: drop
    slug = ''.join(out)
    slug = re.sub(r'-+', '-', slug).strip('-')
    return slug or fallback


def stringify(v) -> str | None:
    """Best-effort conversion of a cell value to a clean string."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    # Numbers: format ints without .0
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def parse(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    categories = []   # list of {slug, label, sort_order}
    dishes = []       # list of {slug, category_slug, name, description, timing, weight, nutrition, serving, sort_order}
    seen_cat_slugs = set()
    seen_dish_slugs: dict[str, int] = {}
    current_cat_slug = None
    cat_sort = 0
    dish_sort_global = 0
    dish_sort_in_cat = 0

    HEADER_NAMES = {'Наименование', 'Меню кухня КОЛЛЕКТИВ ', 'Меню кухня КОЛЛЕКТИВ'}

    for row in ws.iter_rows(values_only=True):
        a = stringify(row[0])
        if not a or a in HEADER_NAMES:
            continue
        b = stringify(row[1]) if len(row) > 1 else None
        c = stringify(row[2]) if len(row) > 2 else None
        d = stringify(row[3]) if len(row) > 3 else None
        e = stringify(row[4]) if len(row) > 4 else None
        f = stringify(row[5]) if len(row) > 5 else None

        is_category = (b is None and c is None and d is None and e is None and f is None)
        if is_category:
            slug = slugify(a, fallback=f'cat-{len(categories)+1}')
            # avoid collision
            base = slug; i = 2
            while slug in seen_cat_slugs:
                slug = f'{base}-{i}'; i += 1
            seen_cat_slugs.add(slug)
            categories.append({'slug': slug, 'label': a, 'sort_order': cat_sort})
            cat_sort += 1
            current_cat_slug = slug
            dish_sort_in_cat = 0
            continue

        if current_cat_slug is None:
            # Dish before any category seen — bucket into "разное"
            current_cat_slug = '_uncategorised'
            if current_cat_slug not in seen_cat_slugs:
                categories.append({'slug': current_cat_slug, 'label': 'Разное', 'sort_order': cat_sort})
                cat_sort += 1
                seen_cat_slugs.add(current_cat_slug)

        slug = slugify(a, fallback=f'dish-{dish_sort_global+1}')
        # dedupe slug
        base = slug; n = seen_dish_slugs.get(base, 0)
        if n:
            slug = f'{base}-{n+1}'
        seen_dish_slugs[base] = n + 1

        dishes.append({
            'slug': slug[:64],
            'category_slug': current_cat_slug,
            'name': a,
            'description': b,
            'timing': c,
            'weight': d,
            'nutrition': e,
            'serving': f,
            'sort_order': dish_sort_in_cat,
        })
        dish_sort_in_cat += 1
        dish_sort_global += 1

    return {'categories': categories, 'dishes': dishes}


def main():
    if len(sys.argv) < 3:
        print("usage: parse-kitchen-xlsx.py <input.xlsx> <output.json>", file=sys.stderr)
        sys.exit(2)
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2])
    data = parse(src)
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"  categories: {len(data['categories'])}")
    print(f"  dishes:     {len(data['dishes'])}")
    print(f"  wrote: {dst}")


if __name__ == '__main__':
    main()
